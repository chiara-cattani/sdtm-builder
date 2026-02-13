"""
Clean Study_Metadata.xlsx by working at the ZIP/XML level to avoid the broken drawing issue.
"""
import os, re, shutil, tempfile
from zipfile import ZipFile

base = r'C:\Users\chiar\OneDrive - Alma Mater Studiorum Università di Bologna\Desktop\Repo\sdtm-builder'
path = os.path.join(base, 'study-pilot', 'metadata', 'Study_Metadata.xlsx')

tmp_dir = tempfile.mkdtemp()
tmp_path = os.path.join(tmp_dir, 'Study_Metadata_clean.xlsx')

with ZipFile(path, 'r') as zin:
    with ZipFile(tmp_path, 'w') as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            
            # Skip any drawing-related files entirely
            if 'drawing' in item.filename.lower():
                print(f"  Removing: {item.filename}")
                continue
            
            # Remove drawing references from .rels files
            if item.filename.endswith('.rels'):
                text = data.decode('utf-8')
                if 'drawing' in text.lower():
                    text = re.sub(r'<Relationship[^>]*[Dd]rawing[^>]*/>', '', text)
                    print(f"  Cleaned drawing refs from: {item.filename}")
                    data = text.encode('utf-8')
            
            # Remove <drawing> elements from worksheet XMLs
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                text = data.decode('utf-8')
                if 'drawing' in text.lower():
                    text = re.sub(r'<drawing[^/]*/>', '', text)
                    text = re.sub(r'<drawing .*?</drawing>', '', text, flags=re.DOTALL)
                    print(f"  Cleaned drawing element from: {item.filename}")
                    data = text.encode('utf-8')
            
            # Remove drawing references from [Content_Types].xml
            if item.filename == '[Content_Types].xml':
                text = data.decode('utf-8')
                if 'drawing' in text.lower():
                    text = re.sub(r'<Override[^>]*[Dd]rawing[^>]*/>', '', text)
                    print(f"  Cleaned drawing from: {item.filename}")
                    data = text.encode('utf-8')
            
            zout.writestr(item, data)

print(f"\nCleaned xlsx written to: {tmp_path}")

# Now try loading with openpyxl
import openpyxl
wb = openpyxl.load_workbook(tmp_path)
print(f"Loaded successfully! Sheets: {wb.sheetnames}")

# === 1. Clean Variables sheet: drop internal notes and USED_IN ===
ws_var = wb['Variables']
headers = [c.value for c in ws_var[1]]
print(f"\nVariables has {len(headers)} columns")

cols_to_delete = []
for idx, h in enumerate(headers, 1):
    if h in ('INTERNALNOTES', 'USED_IN'):
        cols_to_delete.append((idx, h))
        print(f"  Will delete: {h} (col {idx})")

for col_idx, col_name in sorted(cols_to_delete, reverse=True):
    ws_var.delete_cols(col_idx)
    print(f"  Deleted: {col_name}")

# === 2. Clean Meta sheet ===
ws_meta = wb['Meta']
meta_headers = [c.value for c in ws_meta[1]]

for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row):
    for cell in row:
        if cell.value is None:
            continue
        col_h = meta_headers[cell.column - 1] if cell.column <= len(meta_headers) else None
        val = str(cell.value)
        
        if col_h == 'METADATA_DESCRIPTION' and 'SPONSOR' in val:
            cell.value = val.replace('SPONSOR ', '')
            print(f"  Fixed METADATA_DESCRIPTION: '{val}' -> '{cell.value}'")
        
        if col_h == 'STUDY_DESCRIPTION':
            cell.value = "A pilot study demonstrating SDTM domain building with sdtmbuilder."
            print(f"  Fixed STUDY_DESCRIPTION")

# === 3. Clean Standards sheet ===
ws_std = wb['Standards']
std_headers = [c.value for c in ws_std[1]]

for row in ws_std.iter_rows(min_row=2, max_row=ws_std.max_row):
    for cell in row:
        if cell.value is None:
            continue
        col_h = std_headers[cell.column - 1] if cell.column <= len(std_headers) else None
        val = str(cell.value)
        
        if col_h == 'STANDARD_OID' and 'SPONSOR' in val:
            cell.value = val.replace('SPONSOR', 'CDISC')
            print(f"  Fixed STANDARD_OID: '{val}' -> '{cell.value}'")

# Save back to original
wb.save(path)
print(f"\nSaved to: {path}")

# Cleanup
shutil.rmtree(tmp_dir, ignore_errors=True)

# === Verify ===
print("\n=== VERIFICATION ===")

import pandas as pd
xls = pd.ExcelFile(path, engine='openpyxl')

df_var = pd.read_excel(xls, sheet_name='Variables')
print(f"Variables columns ({len(df_var.columns)}): {list(df_var.columns)}")
assert 'INTERNALNOTES' not in df_var.columns, "FAIL: INTERNALNOTES still present!"
assert 'USED_IN' not in df_var.columns, "FAIL: USED_IN still present!"
print("  OK: INTERNALNOTES and USED_IN removed")

df_meta = pd.read_excel(xls, sheet_name='Meta')
for _, row in df_meta.iterrows():
    desc = str(row.get('METADATA_DESCRIPTION', ''))
    assert 'SPONSOR' not in desc, f"FAIL: SPONSOR still in METADATA_DESCRIPTION"
    print(f"  METADATA_DESCRIPTION: {desc}")
    print(f"  STUDY_DESCRIPTION: {row.get('STUDY_DESCRIPTION', '')}")

df_std = pd.read_excel(xls, sheet_name='Standards')
for _, row in df_std.iterrows():
    oid = str(row.get('STANDARD_OID', ''))
    assert 'SPONSOR' not in oid, f"FAIL: SPONSOR still in STANDARD_OID"
    print(f"  STANDARD_OID: {oid}")

# Final scan
print("\n  Final scan for identifiers across all sheets...")
found = False
for sn in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=sn)
    for col in df.columns:
        if re.search(r'\bSPONSOR\b', str(col), re.I):
            print(f"  WARN: Column name '{col}' in sheet {sn}")
            found = True
        vals = df[col].dropna().astype(str)
        matches = vals[vals.str.contains(r'\bSPONSOR\b', case=False, na=False)]
        if len(matches) > 0:
            print(f"  WARN: Values in {sn}.{col}: {matches.values[:3]}")
            found = True

if not found:
    print("  CLEAN: No identifiers found!")

print("\nDONE!")
